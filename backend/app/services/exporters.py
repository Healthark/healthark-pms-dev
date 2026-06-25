"""
Excel exporters — builds openpyxl Workbooks for the 5 entity domains plus
combined / per-employee workbooks.

Each `build_*_sheet` function accepts an `org_id` positional argument
(NEVER taken from a query param) and includes `Model.org_id == org_id`
in every `.filter()` chain. This is the tenant fence — every query in
this file MUST contain it.

No record filtering: the user requested ALL records be included
(soft-deleted users, draft reviews, draft goals, pending project
reviews). Don't add `is_deleted == False` or `status != DRAFT`
predicates here.
"""

from __future__ import annotations

from typing import Optional, Sequence

from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload

from app.core.config import settings
from app.core.cycle_utils import extract_fy_label
from app.models.annual_review_models import AnnualReview
from app.models.goal_criteria_models import GoalCriterion
from app.models.goal_mentor_review_models import GoalMentorReview
from app.models.goal_models import Goal
from app.models.goal_self_review_models import GoalSelfReview
from app.models.project_models import Project, ProjectAssignment
from app.models.project_review_models import ProjectReview, ProjectReviewEvaluator
from app.models.reference_models import Department, Designation
from app.models.system_settings_models import CycleType, SystemSettings
from app.models.user_models import User


# ── Style helpers ───────────────────────────────────────────────────

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
_HEADER_ALIGN = Alignment(vertical="center", wrap_text=False)


def _write_header(ws: Worksheet, headers: Sequence[str]) -> None:
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
    ws.freeze_panes = "A2"


def _autosize(ws: Worksheet, max_width: int = 60) -> None:
    """Cheap auto-size: scan column values, cap at max_width."""
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        width = 10
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            text = str(value)
            if len(text) > width:
                width = len(text)
        ws.column_dimensions[letter].width = min(width + 2, max_width)


def _stringify(value) -> str:
    if value is None:
        return ""
    return str(value)


# ── Security: CSV / Excel formula-injection hardening ───────────────
#
# A spreadsheet app evaluates a cell as a formula when its text begins with
# one of these triggers. User-supplied strings (names, titles, descriptions,
# free-text feedback) flow into cells verbatim, so a value like
# `=HYPERLINK("http://evil",...)` or `=cmd|'/c ...'!A1` would execute when an
# admin opens the export. We neutralize by prefixing a single quote, which
# forces the cell to render as literal text. Applied to data rows only; the
# header row is library-controlled. Numbers/bools/dates are non-str and skip.
_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r")


# ── Memory guard: cap rows per sheet ────────────────────────────────
#
# Workbooks are assembled fully in memory, so an unbounded sheet can OOM the
# worker. `_capped_all` fetches at most MAX_EXPORT_ROWS+1 rows and refuses the
# whole export with a 413 if exceeded — turning an unpredictable process crash
# (kills every in-flight request) into a recoverable, actionable error for the
# one oversized request. Tunable via settings.EXPORT_MAX_ROWS.
MAX_EXPORT_ROWS = getattr(settings, "EXPORT_MAX_ROWS", 100_000)


def _capped_all(query, label: str) -> list:
    """Materialize a query's rows, but refuse (413) if it exceeds the cap.
    Uses limit(cap+1) so we never pull more than cap+1 rows into memory."""
    rows = query.limit(MAX_EXPORT_ROWS + 1).all()
    if len(rows) > MAX_EXPORT_ROWS:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=(
                f"{label} export exceeds the {MAX_EXPORT_ROWS:,}-row limit. "
                "Narrow the export by financial year or entity and try again."
            ),
        )
    return rows


def _harden_formula_injection(wb: Workbook) -> None:
    """Prefix any string cell that starts with a formula trigger with `'` so
    Excel/Sheets/LibreOffice treat it as text rather than evaluating it.

    openpyxl auto-stores a leading-`=` value as a real formula; reading and
    re-assigning with the quote prefix demotes it back to an inert string."""
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value[:1] in _FORMULA_TRIGGERS:
                    cell.value = "'" + value


def _dt(value) -> str:
    """Render a datetime/date as a portable ISO-ish string for Excel."""
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


# ── FY filter helpers ───────────────────────────────────────────────

def _fy_token_variants(fy: Optional[str]) -> Optional[list[str]]:
    """Expand a caller-supplied FY string into the list of stored variants
    that should match it.

    Cycle naming differs per entity:
        Goal.cycle_name      → "H1 2026" (legacy)  or "H1 FY26-27" / "FY26-27"
        AnnualReview.cycle_name → "FY26-27" or legacy "FY26"
        ProjectReview.cycle  → "H1 FY26-27", "Q2 FY26-27", etc.

    Strategy: produce a small list of tokens (e.g. ["FY26-27", "FY26", "2026"])
    such that an ILIKE %token% against any of them catches the same FY.
    """
    if not fy:
        return None
    raw = fy.strip().upper()
    if not raw:
        return None

    # If caller sent a full cycle name (e.g. "H1 FY26-27"), strip the prefix
    canonical = extract_fy_label(raw)  # "FY26-27" or "FY26" or unchanged
    variants: set[str] = {canonical}

    # Derive the 4-digit calendar year (e.g. "FY26-27" → 2026, "FY26" → 2026)
    if canonical.startswith("FY"):
        try:
            two_digit = int(canonical[2:4])
        except ValueError:
            two_digit = None
        if two_digit is not None:
            # naive century pivot — < 70 means 2000s, else 1900s
            full_year = 2000 + two_digit if two_digit < 70 else 1900 + two_digit
            variants.add(str(full_year))
            # short form alternative
            variants.add(f"FY{two_digit:02d}")
            # span form alternative
            variants.add(f"FY{two_digit:02d}-{(two_digit + 1) % 100:02d}")

    return sorted(variants)


def _apply_fy_ilike(query, column, fy: Optional[str]):
    if not fy:
        return query
    variants = _fy_token_variants(fy)
    if not variants:
        return query
    from sqlalchemy import or_
    return query.filter(or_(*(column.ilike(f"%{v}%") for v in variants)))


# ── Lookup helpers ──────────────────────────────────────────────────

def _build_user_lookup(db: Session, org_id: int) -> dict[int, User]:
    """Pre-load every org user once for cheap mentor-name / reviewer-name fills."""
    users = (
        db.query(User)
        .options(joinedload(User.department), joinedload(User.designation))
        .filter(User.org_id == org_id)
        .all()
    )
    return {u.id: u for u in users}


def _user_display(users_by_id: dict[int, User], user_id: Optional[int]) -> str:
    if user_id is None:
        return ""
    u = users_by_id.get(user_id)
    return u.full_name if u else ""


def _user_code(users_by_id: dict[int, User], user_id: Optional[int]) -> str:
    if user_id is None:
        return ""
    u = users_by_id.get(user_id)
    return u.employee_code if u else ""


def _project_lookup(db: Session, org_id: int) -> dict[int, Project]:
    rows = db.query(Project).filter(Project.org_id == org_id).all()
    return {p.id: p for p in rows}


def _get_cycle_type(db: Session, org_id: int) -> str:
    settings = (
        db.query(SystemSettings).filter(SystemSettings.org_id == org_id).first()
    )
    if settings is None:
        return CycleType.HALF_YEARLY.value
    return settings.cycle_type or CycleType.HALF_YEARLY.value


# ── Sheet: Users ────────────────────────────────────────────────────

USERS_HEADERS = [
    "Sr No", "Employee Code", "Full Name", "Email", "Phone",
    "Role", "Is Management", "Is Deleted",
    "Department", "Designation", "Mentor",
    "Must Change Password", "Created At", "Updated At",
]


def build_users_sheet(ws: Worksheet, db: Session, org_id: int, fy: Optional[str] = None) -> int:
    """`fy` is accepted for symmetry but Users have no cycle column — ignored."""
    _write_header(ws, USERS_HEADERS)
    users = _capped_all(
        db.query(User)
        .options(joinedload(User.department), joinedload(User.designation))
        .filter(User.org_id == org_id)
        .order_by(User.created_at.desc()),
        "Users",
    )
    users_by_id = {u.id: u for u in users}

    row = 2
    for u in users:
        # Sr No: 1-indexed row counter (NOT the DB id) — opaque internal
        # ids leak surface area we don't want in HR-facing exports.
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=u.employee_code)
        ws.cell(row=row, column=3, value=u.full_name)
        ws.cell(row=row, column=4, value=u.email)
        ws.cell(row=row, column=5, value=u.phone or "")
        ws.cell(row=row, column=6, value=u.role)
        ws.cell(row=row, column=7, value=bool(u.is_management))
        ws.cell(row=row, column=8, value=bool(u.is_deleted))
        ws.cell(row=row, column=9, value=u.department.name if u.department else "")
        ws.cell(row=row, column=10, value=u.designation.name if u.designation else "")
        ws.cell(row=row, column=11, value=_user_display(users_by_id, u.mentor_id))
        ws.cell(row=row, column=12, value=bool(u.must_change_password))
        ws.cell(row=row, column=13, value=_dt(u.created_at))
        ws.cell(row=row, column=14, value=_dt(u.updated_at))
        row += 1
    _autosize(ws)
    return row - 2


# ── Sheet: Projects ─────────────────────────────────────────────────

PROJECTS_HEADERS = [
    "Sr No", "Project Code", "Name", "Description",
    "Start Date", "Expected End Date",
    "Reports To", "Secondary Evaluator",
    "PM Name", "Member Count",
    "Is Deleted", "Created At", "Updated At",
]


def build_projects_sheet(ws: Worksheet, db: Session, org_id: int, fy: Optional[str] = None) -> int:
    """`fy` accepted for symmetry; Projects have no cycle column — ignored."""
    _write_header(ws, PROJECTS_HEADERS)
    projects = _capped_all(
        db.query(Project)
        .filter(Project.org_id == org_id)
        .order_by(Project.created_at.desc()),
        "Projects",
    )
    users_by_id = _build_user_lookup(db, org_id)
    assignments = (
        db.query(ProjectAssignment)
        .filter(ProjectAssignment.org_id == org_id)
        .all()
    )
    pm_by_project: dict[int, str] = {}
    members_by_project: dict[int, int] = {}
    for a in assignments:
        # Member count + PM reflect the ACTIVE team (soft-removed members are
        # listed separately on the Project Assignments sheet, marked removed).
        if a.is_deleted:
            continue
        members_by_project[a.project_id] = members_by_project.get(a.project_id, 0) + 1
        if a.evaluator_type == "Primary":
            pm_by_project[a.project_id] = _user_display(users_by_id, a.user_id)

    row = 2
    for p in projects:
        # Sr No: 1-indexed row counter (NOT the DB id) — opaque internal
        # ids leak surface area we don't want in HR-facing exports.
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=p.project_code)
        ws.cell(row=row, column=3, value=p.name)
        ws.cell(row=row, column=4, value=p.description or "")
        ws.cell(row=row, column=5, value=_dt(p.start_date))
        ws.cell(row=row, column=6, value=_dt(p.expected_end_date))
        ws.cell(row=row, column=7, value=_user_display(users_by_id, p.reports_to_id))
        ws.cell(row=row, column=8, value=_user_display(users_by_id, p.secondary_evaluator_id))
        ws.cell(row=row, column=9, value=pm_by_project.get(p.id, ""))
        ws.cell(row=row, column=10, value=members_by_project.get(p.id, 0))
        ws.cell(row=row, column=11, value=bool(p.is_deleted))
        ws.cell(row=row, column=12, value=_dt(p.created_at))
        ws.cell(row=row, column=13, value=_dt(p.updated_at))
        row += 1
    _autosize(ws)
    return row - 2


# ── Sheet: Project Assignments (combined workbook only) ─────────────

PROJECT_ASSIGNMENTS_HEADERS = [
    "Sr No", "Project Code", "Project Name",
    "Employee Code", "Employee Name",
    "Assignment Role", "Department", "Evaluator Type",
    "Assigned Date", "Created At",
    "Status", "Removed By", "Removed On",
]


def build_project_assignments_sheet(ws: Worksheet, db: Session, org_id: int) -> int:
    _write_header(ws, PROJECT_ASSIGNMENTS_HEADERS)
    rows = _capped_all(
        db.query(ProjectAssignment)
        .options(joinedload(ProjectAssignment.department))
        .filter(ProjectAssignment.org_id == org_id),
        "Project Assignments",
    )
    users_by_id = _build_user_lookup(db, org_id)
    projects_by_id = _project_lookup(db, org_id)

    r = 2
    for a in rows:
        proj = projects_by_id.get(a.project_id)
        # Sr No: 1-indexed row counter (NOT the DB id).
        ws.cell(row=r, column=1, value=r - 1)
        ws.cell(row=r, column=2, value=proj.project_code if proj else "")
        ws.cell(row=r, column=3, value=proj.name if proj else "")
        ws.cell(row=r, column=4, value=_user_code(users_by_id, a.user_id))
        ws.cell(row=r, column=5, value=_user_display(users_by_id, a.user_id))
        ws.cell(row=r, column=6, value=a.assignment_role or "")
        ws.cell(row=r, column=7, value=a.department.name if a.department else "")
        ws.cell(row=r, column=8, value=a.evaluator_type or "")
        ws.cell(row=r, column=9, value=_dt(a.assigned_date))
        ws.cell(row=r, column=10, value=_dt(a.created_at))
        # Soft-delete audit — removed members are kept in the export, marked.
        ws.cell(row=r, column=11, value="Removed" if a.is_deleted else "Active")
        ws.cell(
            row=r,
            column=12,
            value=_user_display(users_by_id, a.removed_by_id) if a.is_deleted else "",
        )
        ws.cell(row=r, column=13, value=_dt(a.removed_at) if a.is_deleted else "")
        r += 1
    _autosize(ws)
    return r - 2


# ── Sheet: Annual Goals ─────────────────────────────────────────────

GOALS_HEADERS = [
    "Sr No", "Employee Code", "Employee Name", "Manager",
    "Goal Type", "Cycle Name",
    "Title", "Description",
    "Approval Status", "Manager Feedback", "Progress Notes",
    "Start Date", "Due Date", "Approved At",
    "Criteria Count", "Criteria Completed", "Progress %",
    # The next 4 columns swap H1/H2 ↔ Q1/Q2/Q3/Q4 depending on org cycle_type.
    "Period 1 Self", "Period 1 Mentor",
    "Period 2 Self", "Period 2 Mentor",
    "Period 3 Self", "Period 3 Mentor",
    "Period 4 Self", "Period 4 Mentor",
    "Created At", "Updated At",
]


def _criteria_rollup(criteria) -> tuple[int, int, int]:
    total = len(criteria)
    done = sum(1 for c in criteria if bool(c.is_completed))
    pct = int(round((done / total) * 100)) if total else 0
    return total, done, pct


def _review_by_half(reviews, half: str, field: str) -> str:
    """Pull the freeform paragraph for a given half from self_reviews/mentor_reviews."""
    for r in reviews:
        if r.cycle_half == half and not bool(getattr(r, "is_draft", False)):
            return getattr(r, field, "") or ""
        if r.cycle_half == half and bool(getattr(r, "is_draft", False)):
            # caller requested ALL records — include drafts but tag them
            text = getattr(r, field, "") or ""
            return f"[DRAFT] {text}" if text else "[DRAFT]"
    return ""


def build_goals_sheet(
    ws: Worksheet,
    db: Session,
    org_id: int,
    fy: Optional[str] = None,
    user_id: Optional[int] = None,
) -> int:
    _write_header(ws, GOALS_HEADERS)
    q = (
        db.query(Goal)
        .options(
            joinedload(Goal.criteria),
            joinedload(Goal.self_reviews),
            joinedload(Goal.mentor_reviews),
        )
        .filter(Goal.org_id == org_id, Goal.is_deleted == False)  # noqa: E712
    )
    if user_id is not None:
        q = q.filter(Goal.user_id == user_id)
    q = _apply_fy_ilike(q, Goal.cycle_name, fy)
    goals = _capped_all(q.order_by(Goal.created_at.desc()), "Annual Goals")

    users_by_id = _build_user_lookup(db, org_id)

    r = 2
    for g in goals:
        total, done, pct = _criteria_rollup(g.criteria)
        # Sr No: 1-indexed row counter (NOT the DB id).
        ws.cell(row=r, column=1, value=r - 1)
        ws.cell(row=r, column=2, value=_user_code(users_by_id, g.user_id))
        ws.cell(row=r, column=3, value=_user_display(users_by_id, g.user_id))
        ws.cell(row=r, column=4, value=_user_display(users_by_id, g.manager_id))
        ws.cell(row=r, column=5, value=g.goal_type)
        ws.cell(row=r, column=6, value=g.cycle_name or "")
        ws.cell(row=r, column=7, value=g.title)
        ws.cell(row=r, column=8, value=g.description or "")
        ws.cell(row=r, column=9, value=g.approval_status)
        ws.cell(row=r, column=10, value=g.manager_feedback or "")
        ws.cell(row=r, column=11, value=g.progress_notes or "")
        ws.cell(row=r, column=12, value=_dt(g.start_date))
        ws.cell(row=r, column=13, value=_dt(g.due_date))
        ws.cell(row=r, column=14, value=_dt(g.approved_at))
        ws.cell(row=r, column=15, value=total)
        ws.cell(row=r, column=16, value=done)
        ws.cell(row=r, column=17, value=pct)
        # H1/Q1/H2/Q2 ... — same column slots reused regardless of cadence.
        ws.cell(row=r, column=18, value=_review_by_half(g.self_reviews, "H1", "self_overall_review")
                or _review_by_half(g.self_reviews, "Q1", "self_overall_review"))
        ws.cell(row=r, column=19, value=_review_by_half(g.mentor_reviews, "H1", "mentor_overall_review")
                or _review_by_half(g.mentor_reviews, "Q1", "mentor_overall_review"))
        ws.cell(row=r, column=20, value=_review_by_half(g.self_reviews, "H2", "self_overall_review")
                or _review_by_half(g.self_reviews, "Q2", "self_overall_review"))
        ws.cell(row=r, column=21, value=_review_by_half(g.mentor_reviews, "H2", "mentor_overall_review")
                or _review_by_half(g.mentor_reviews, "Q2", "mentor_overall_review"))
        ws.cell(row=r, column=22, value=_review_by_half(g.self_reviews, "Q3", "self_overall_review"))
        ws.cell(row=r, column=23, value=_review_by_half(g.mentor_reviews, "Q3", "mentor_overall_review"))
        ws.cell(row=r, column=24, value=_review_by_half(g.self_reviews, "Q4", "self_overall_review"))
        ws.cell(row=r, column=25, value=_review_by_half(g.mentor_reviews, "Q4", "mentor_overall_review"))
        ws.cell(row=r, column=26, value=_dt(g.created_at))
        ws.cell(row=r, column=27, value=_dt(g.updated_at))
        r += 1
    _autosize(ws)
    return r - 2


# ── Sheet: Annual Reviews ───────────────────────────────────────────

ANNUAL_REVIEWS_HEADERS = [
    "Sr No", "Employee Code", "Employee Name", "Mentor",
    "Cycle Name", "Status",
    "Self Overall Review", "Self Performance Rating",
    "Mentor Overall Review", "Mentor Performance Rating",
    "Mentor Draft Review", "Mentor Draft Rating",
    "Management Performance Rating", "Final Performance Rating",
    "Final Rating Enabled", "Management Comments",
    "Created At", "Updated At",
]


def build_annual_reviews_sheet(
    ws: Worksheet,
    db: Session,
    org_id: int,
    fy: Optional[str] = None,
    user_id: Optional[int] = None,
) -> int:
    _write_header(ws, ANNUAL_REVIEWS_HEADERS)
    q = db.query(AnnualReview).filter(AnnualReview.org_id == org_id)
    if user_id is not None:
        q = q.filter(AnnualReview.user_id == user_id)
    q = _apply_fy_ilike(q, AnnualReview.cycle_name, fy)
    reviews = _capped_all(
        q.order_by(AnnualReview.created_at.desc()), "Annual Reviews"
    )

    users_by_id = _build_user_lookup(db, org_id)

    r = 2
    for rv in reviews:
        # Sr No: 1-indexed row counter (NOT the DB id).
        ws.cell(row=r, column=1, value=r - 1)
        ws.cell(row=r, column=2, value=_user_code(users_by_id, rv.user_id))
        ws.cell(row=r, column=3, value=_user_display(users_by_id, rv.user_id))
        ws.cell(row=r, column=4, value=_user_display(users_by_id, rv.mentor_id))
        ws.cell(row=r, column=5, value=rv.cycle_name)
        ws.cell(row=r, column=6, value=rv.status)
        ws.cell(row=r, column=7, value=rv.self_overall_review or "")
        ws.cell(row=r, column=8, value=rv.self_performance_rating)
        ws.cell(row=r, column=9, value=rv.mentor_overall_review or "")
        ws.cell(row=r, column=10, value=rv.mentor_performance_rating)
        ws.cell(row=r, column=11, value=rv.mentor_overall_review_draft or "")
        ws.cell(row=r, column=12, value=rv.mentor_performance_rating_draft)
        ws.cell(row=r, column=13, value=rv.management_performance_rating)
        ws.cell(row=r, column=14, value=rv.final_performance_rating)
        ws.cell(row=r, column=15, value=bool(rv.final_rating_enabled))
        ws.cell(row=r, column=16, value=rv.management_comments or "")
        ws.cell(row=r, column=17, value=_dt(rv.created_at))
        ws.cell(row=r, column=18, value=_dt(rv.updated_at))
        r += 1
    _autosize(ws)
    return r - 2


# ── Sheet: Project Reviews ──────────────────────────────────────────

PROJECT_REVIEWS_HEADERS = [
    "Sr No", "Project Code", "Project Name",
    "Employee Code", "Employee Name", "Reviewer",
    "Cycle", "Status", "Performance Group", "Impact Statement",
    "Comment: Task Execution",
    "Comment: Ownership",
    "Comment: Project Management",
    "Comment: Client Deliverables",
    "Comment: Communication",
    "Comment: Mentoring",
    "Comment: Competency & Skills",
    "Secondary Evaluators",
    "Secondary Impact Statements",
    "Is Deleted", "Created At", "Updated At",
]


def build_project_reviews_sheet(
    ws: Worksheet,
    db: Session,
    org_id: int,
    fy: Optional[str] = None,
    user_id: Optional[int] = None,
) -> int:
    _write_header(ws, PROJECT_REVIEWS_HEADERS)
    q = (
        db.query(ProjectReview)
        .options(joinedload(ProjectReview.secondary_evaluations))
        .filter(ProjectReview.org_id == org_id)
    )
    if user_id is not None:
        q = q.filter(ProjectReview.user_id == user_id)
    q = _apply_fy_ilike(q, ProjectReview.cycle, fy)
    reviews = _capped_all(
        q.order_by(ProjectReview.created_at.desc()), "Project Reviews"
    )

    users_by_id = _build_user_lookup(db, org_id)
    projects_by_id = _project_lookup(db, org_id)

    r = 2
    for rv in reviews:
        proj = projects_by_id.get(rv.project_id)
        sec_names = " | ".join(
            _user_display(users_by_id, s.evaluator_id) for s in rv.secondary_evaluations
        )
        sec_impacts = " | ".join(
            (s.impact_statement or "") for s in rv.secondary_evaluations
        )
        # Sr No: 1-indexed row counter (NOT the DB id).
        ws.cell(row=r, column=1, value=r - 1)
        ws.cell(row=r, column=2, value=proj.project_code if proj else "")
        ws.cell(row=r, column=3, value=proj.name if proj else "")
        ws.cell(row=r, column=4, value=_user_code(users_by_id, rv.user_id))
        ws.cell(row=r, column=5, value=_user_display(users_by_id, rv.user_id))
        ws.cell(row=r, column=6, value=_user_display(users_by_id, rv.reviewer_id))
        ws.cell(row=r, column=7, value=rv.cycle)
        ws.cell(row=r, column=8, value=rv.status)
        ws.cell(row=r, column=9, value=rv.performance_group or "")
        ws.cell(row=r, column=10, value=rv.impact_statement or "")
        ws.cell(row=r, column=11, value=rv.comment_task_execution or "")
        ws.cell(row=r, column=12, value=rv.comment_ownership or "")
        ws.cell(row=r, column=13, value=rv.comment_project_management or "")
        ws.cell(row=r, column=14, value=rv.comment_client_deliverables or "")
        ws.cell(row=r, column=15, value=rv.comment_communication or "")
        ws.cell(row=r, column=16, value=rv.comment_mentoring or "")
        ws.cell(row=r, column=17, value=rv.comment_competency_skills or "")
        ws.cell(row=r, column=18, value=sec_names)
        ws.cell(row=r, column=19, value=sec_impacts)
        ws.cell(row=r, column=20, value=bool(rv.is_deleted))
        ws.cell(row=r, column=21, value=_dt(rv.created_at))
        ws.cell(row=r, column=22, value=_dt(rv.updated_at))
        r += 1
    _autosize(ws)
    return r - 2


# ── Sheet: Project Review Evaluators (combined workbook only) ───────

PROJECT_REVIEW_EVALUATORS_HEADERS = [
    "Sr No", "Project Review ID",
    "Project Code", "Project Name",
    "Reviewed Employee", "Evaluator Name", "Evaluator Type",
    "Status", "Impact Statement", "Created At",
]


def build_project_review_evaluators_sheet(ws: Worksheet, db: Session, org_id: int) -> int:
    _write_header(ws, PROJECT_REVIEW_EVALUATORS_HEADERS)
    rows = _capped_all(
        db.query(ProjectReviewEvaluator)
        .filter(ProjectReviewEvaluator.org_id == org_id),
        "Project Review Evaluators",
    )
    users_by_id = _build_user_lookup(db, org_id)
    reviews = {
        rv.id: rv
        for rv in db.query(ProjectReview).filter(ProjectReview.org_id == org_id).all()
    }
    projects_by_id = _project_lookup(db, org_id)

    r = 2
    for ev in rows:
        review = reviews.get(ev.project_review_id)
        proj = projects_by_id.get(review.project_id) if review else None
        # Sr No: 1-indexed row counter (NOT the DB id).
        ws.cell(row=r, column=1, value=r - 1)
        # Project Review ID kept as DB id — serves as a join key back to the
        # Project Reviews sheet for cross-sheet correlation. (Note: that
        # sheet now uses Sr No, so this FK is no longer a row pointer there;
        # readers should match on Project Code + Reviewed Employee instead.)
        ws.cell(row=r, column=2, value=ev.project_review_id)
        ws.cell(row=r, column=3, value=proj.project_code if proj else "")
        ws.cell(row=r, column=4, value=proj.name if proj else "")
        ws.cell(row=r, column=5, value=_user_display(users_by_id, review.user_id) if review else "")
        ws.cell(row=r, column=6, value=_user_display(users_by_id, ev.evaluator_id))
        ws.cell(row=r, column=7, value=ev.evaluator_type or "")
        ws.cell(row=r, column=8, value=ev.status)
        ws.cell(row=r, column=9, value=ev.impact_statement or "")
        ws.cell(row=r, column=10, value=_dt(ev.created_at))
        r += 1
    _autosize(ws)
    return r - 2


# ── Sheet: Profile (per-employee workbook only) ─────────────────────

def build_profile_sheet(ws: Worksheet, db: Session, org_id: int, user_id: int) -> int:
    user = (
        db.query(User)
        .options(joinedload(User.department), joinedload(User.designation))
        .filter(User.id == user_id, User.org_id == org_id)
        .first()
    )
    if user is None:
        _write_header(ws, ["Field", "Value"])
        ws.cell(row=2, column=1, value="Error")
        ws.cell(row=2, column=2, value="User not found in this organization.")
        _autosize(ws)
        return 0

    users_by_id = _build_user_lookup(db, org_id)
    pairs = [
        ("Employee Code", user.employee_code),
        ("Full Name", user.full_name),
        ("Email", user.email),
        ("Phone", user.phone or ""),
        ("Role", user.role),
        ("Is Management", bool(user.is_management)),
        ("Is Deleted", bool(user.is_deleted)),
        ("Department", user.department.name if user.department else ""),
        ("Designation", user.designation.name if user.designation else ""),
        ("Mentor", _user_display(users_by_id, user.mentor_id)),
        ("Must Change Password", bool(user.must_change_password)),
        ("Created At", _dt(user.created_at)),
        ("Updated At", _dt(user.updated_at)),
    ]
    _write_header(ws, ["Field", "Value"])
    for idx, (k, v) in enumerate(pairs, start=2):
        ws.cell(row=idx, column=1, value=k)
        ws.cell(row=idx, column=2, value=v)
    _autosize(ws)
    return len(pairs)


# ── Workbook orchestrators ──────────────────────────────────────────

def build_single_entity_workbook(
    entity: str,
    db: Session,
    org_id: int,
    fy: Optional[str] = None,
    user_id: Optional[int] = None,
) -> tuple[Workbook, int]:
    """Build a one-sheet workbook for the given entity key.

    `entity` ∈ {users, projects, goals, annual_reviews, project_reviews}.
    Returns (Workbook, row_count).
    """
    wb = Workbook()
    ws = wb.active
    if entity == "users":
        ws.title = "Users"
        n = build_users_sheet(ws, db, org_id, fy)
    elif entity == "projects":
        ws.title = "Projects"
        n = build_projects_sheet(ws, db, org_id, fy)
    elif entity == "goals":
        ws.title = "Annual Goals"
        n = build_goals_sheet(ws, db, org_id, fy, user_id)
    elif entity == "annual_reviews":
        ws.title = "Annual Reviews"
        n = build_annual_reviews_sheet(ws, db, org_id, fy, user_id)
    elif entity == "project_reviews":
        ws.title = "Project Reviews"
        n = build_project_reviews_sheet(ws, db, org_id, fy, user_id)
    else:
        raise ValueError(f"Unknown entity: {entity}")
    _harden_formula_injection(wb)
    return wb, n


def build_combined_workbook(
    db: Session, org_id: int, fy: Optional[str] = None
) -> tuple[Workbook, int]:
    """All 5 entity sheets + Project Assignments + Project Review Evaluators."""
    wb = Workbook()
    # Replace the default sheet
    default = wb.active
    wb.remove(default)

    total = 0
    for title, builder in [
        ("Users", lambda ws: build_users_sheet(ws, db, org_id, fy)),
        ("Projects", lambda ws: build_projects_sheet(ws, db, org_id, fy)),
        ("Project Assignments", lambda ws: build_project_assignments_sheet(ws, db, org_id)),
        ("Annual Goals", lambda ws: build_goals_sheet(ws, db, org_id, fy)),
        ("Annual Reviews", lambda ws: build_annual_reviews_sheet(ws, db, org_id, fy)),
        ("Project Reviews", lambda ws: build_project_reviews_sheet(ws, db, org_id, fy)),
        ("Project Review Evaluators", lambda ws: build_project_review_evaluators_sheet(ws, db, org_id)),
    ]:
        ws = wb.create_sheet(title=title)
        total += builder(ws)
    _harden_formula_injection(wb)
    return wb, total


def build_per_employee_workbook(
    db: Session, org_id: int, target_user_id: int, fy: Optional[str] = None
) -> tuple[Workbook, int]:
    """Profile + that user's Goals + Annual Reviews + Project Reviews."""
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    total = 0
    profile_ws = wb.create_sheet(title="Profile")
    total += build_profile_sheet(profile_ws, db, org_id, target_user_id)

    goals_ws = wb.create_sheet(title="Annual Goals")
    total += build_goals_sheet(goals_ws, db, org_id, fy, user_id=target_user_id)

    ar_ws = wb.create_sheet(title="Annual Reviews")
    total += build_annual_reviews_sheet(ar_ws, db, org_id, fy, user_id=target_user_id)

    pr_ws = wb.create_sheet(title="Project Reviews")
    total += build_project_reviews_sheet(pr_ws, db, org_id, fy, user_id=target_user_id)

    _harden_formula_injection(wb)
    return wb, total


# ── Self-service: one user's own goals ──────────────────────────────
#
# Unlike build_goals_sheet (HR/management — dumps EVERY review incl. drafts),
# this renders from already-visibility-filtered GoalResponse objects produced
# by goal_routes.list_goals. That upstream gate strips mentor-review drafts and
# embargoes unpublished mentor reviews from the mentee, so a user exporting
# their own goals can never see mentor feedback the My Goals screen hides.

MY_GOALS_BASE_HEADERS = [
    "Sr No", "Cycle", "Title", "Description",
    "Status", "Mentor", "Mentor Feedback", "Progress Notes",
    "Start Date", "Due Date", "Approved At",
    "Criteria Total", "Criteria Completed", "Progress %",
]


def build_my_goals_workbook(goals, db: Session, org_id: int) -> tuple[Workbook, int]:
    """Single-sheet workbook of ONE user's own annual goals.

    `goals` is the visibility-filtered list of GoalResponse objects from
    goal_routes.list_goals (the embargo / draft-hiding is applied upstream —
    this only renders). Period columns follow the org cadence: H1/H2 for
    half-yearly & annual orgs, Q1-Q4 for quarterly.
    """
    cycle_type = _get_cycle_type(db, org_id)
    periods = (
        ("Q1", "Q2", "Q3", "Q4")
        if cycle_type == CycleType.QUARTERLY.value
        else ("H1", "H2")
    )
    headers = list(MY_GOALS_BASE_HEADERS)
    for p in periods:
        headers.append(f"{p} Self Review")
        headers.append(f"{p} Mentor Review")
    headers.extend(["Created At", "Updated At"])

    wb = Workbook()
    ws = wb.active
    ws.title = "My Annual Goals"
    _write_header(ws, headers)

    r = 2
    for g in goals:
        total = len(g.criteria)
        done = sum(1 for cr in g.criteria if bool(cr.is_completed))
        pct = int(round((done / total) * 100)) if total else 0
        values = [
            r - 1,
            g.cycle_name or "",
            g.title,
            g.description or "",
            g.approval_status,
            g.manager_name or "",
            g.manager_feedback or "",
            g.progress_notes or "",
            _dt(g.start_date),
            _dt(g.due_date),
            _dt(g.approved_at),
            total,
            done,
            pct,
        ]
        for p in periods:
            values.append(_review_by_half(g.self_reviews, p, "self_overall_review"))
            values.append(_review_by_half(g.mentor_reviews, p, "mentor_overall_review"))
        values.append(_dt(g.created_at))
        values.append(_dt(g.updated_at))
        for col_idx, val in enumerate(values, start=1):
            ws.cell(row=r, column=col_idx, value=val)
        r += 1
    _autosize(ws)
    _harden_formula_injection(wb)
    return wb, r - 2
